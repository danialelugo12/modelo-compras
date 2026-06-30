"""
Módulo de exportación Excel.
Genera el archivo final con 6 hojas formateadas.
"""

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import CONFIG


# ============================================================================
# ESTILOS
# ============================================================================
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill("solid", start_color="8B0000")
FONT_NORMAL = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="8B0000")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

FILL_ESTADO = {
    "CRÍTICO": PatternFill("solid", start_color="F8CBAD"),
    "COMPRAR": PatternFill("solid", start_color="FFE699"),
    "OK": PatternFill("solid", start_color="C6EFCE"),
    "SOBRESTOCK": PatternFill("solid", start_color="D9D9D9"),
}
FILL_ABC = {
    "A": PatternFill("solid", start_color="C6E0B4"),
    "B": PatternFill("solid", start_color="FFEB9C"),
    "C": PatternFill("solid", start_color="F4B084"),
}

# Formatos numéricos por nombre de columna
FORMATOS = {
    "PRECIO AILOO": "$#,##0;($#,##0);-",
    "ULTIMO PRECIO NETO": "$#,##0;($#,##0);-",
    "COSTO COMPRA AJUSTADA": "$#,##0;($#,##0);-",
    "VALOR EXCESO": "$#,##0;($#,##0);-",
    "COSTO UNITARIO": "$#,##0;($#,##0);-",
    "% PARTICIPACION": "0.00%",
    "% ACUMULADO": "0.00%",
    "DEMANDA MENSUAL": "0.00",
    "COBERTURA (MESES)": "0.00",
    "COBERTURA ACTUAL": "0.00",
    "COBERTURA OBJETIVO": "0.00",
    "COBERTURA OBJ": "0.00",
    "CANT SUGERIDA BASE": "0.00",
    "EXCESO UNIDADES": "0.00",
    "VENTAS 6M": "#,##0",
    "FALLIDAS 6M": "#,##0",
    "DEMANDA TOTAL": "#,##0",
    "STOCK ACTUAL": "#,##0",
    "STOCK": "#,##0",
    "EN TRÁNSITO": "#,##0",
    "MESES A COMPRAR": "#,##0",
    "COMPRA AJUSTADA": "#,##0",
    "VECES_SOLICITADO": "#,##0",
}

ANCHOS = {
    "ID": 10,
    "MARCA": 12,
    "PRODUCTO": 50,
    "SKU": 14,
    "PRECIO AILOO": 12,
    "VENTAS 6M": 10,
    "FALLIDAS 6M": 11,
    "DEMANDA TOTAL": 12,
    "DEMANDA MENSUAL": 14,
    "% PARTICIPACION": 13,
    "% ACUMULADO": 12,
    "ABC": 6,
    "STOCK ACTUAL": 12,
    "STOCK": 10,
    "EN TRÁNSITO": 11,
    "COBERTURA (MESES)": 14,
    "COBERTURA ACTUAL": 14,
    "TIPO ABASTECIMIENTO": 16,
    "COBERTURA OBJETIVO": 15,
    "COBERTURA OBJ": 12,
    "CANT SUGERIDA BASE": 16,
    "ESTADO INVENTARIO": 16,
    "ESTADO": 12,
    "ULTIMO PRECIO NETO": 14,
    "COSTO UNITARIO": 13,
    "ULTIMA FECHA COMPRA": 16,
    "ULTIMO PROVEEDOR": 14,
    "MESES A COMPRAR": 13,
    "COMPRA AJUSTADA": 14,
    "COSTO COMPRA AJUSTADA": 17,
    "EXCESO UNIDADES": 14,
    "VALOR EXCESO": 14,
    "OBSERVACION": 45,
    "PRODUCTO_DESCRIPCION": 60,
    "VECES_SOLICITADO": 18,
}


# ============================================================================
# HELPERS
# ============================================================================
def escribir_fila_dato(ws, fila, col, valor, col_name):
    """Escribe una celda con formato según el nombre de columna."""
    if isinstance(valor, pd.Timestamp):
        valor = valor.strftime("%Y-%m-%d")
    elif isinstance(valor, float) and np.isnan(valor):
        valor = None

    c = ws.cell(row=fila, column=col, value=valor)
    c.font = FONT_NORMAL
    c.border = BORDER_THIN

    if col_name == "ABC":
        c.fill = FILL_ABC.get(valor, PatternFill())
        c.alignment = ALIGN_CENTER
    elif col_name in ("ESTADO INVENTARIO", "ESTADO"):
        c.fill = FILL_ESTADO.get(valor, PatternFill())
        c.alignment = ALIGN_CENTER
    elif col_name in ("ID", "SKU", "MARCA", "TIPO ABASTECIMIENTO",
                       "ULTIMO PROVEEDOR", "ULTIMA FECHA COMPRA"):
        c.alignment = ALIGN_CENTER
    elif col_name == "PRODUCTO":
        c.alignment = ALIGN_LEFT
    else:
        c.alignment = ALIGN_RIGHT

    if col_name in FORMATOS:
        c.number_format = FORMATOS[col_name]
    return c


def escribir_tabla(ws, df, fila_inicio_header=1, fila_inicio_datos=2):
    """Escribe un DataFrame como tabla con headers + datos formateados."""
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=fila_inicio_header, column=j, value=col)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN

    for i, row in enumerate(df.itertuples(index=False), start=fila_inicio_datos):
        for j, val in enumerate(row, start=1):
            escribir_fila_dato(ws, i, j, val, df.columns[j - 1])

    for j, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = ANCHOS.get(col, 12)


# ============================================================================
# HOJAS INDIVIDUALES
# ============================================================================
def hoja_resumen(wb: Workbook, modelo: pd.DataFrame, ranking_sin_id: pd.DataFrame):
    ws = wb.create_sheet("1. Resumen Ejecutivo")
    ws.merge_cells("A1:B1")
    ws["A1"] = "INDICADOR MODELO DE COMPRAS - RESUMEN EJECUTIVO"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center")

    modelo = modelo.copy()
    modelo["VALOR_STOCK"] = modelo["STOCK_ACTUAL"] * modelo["ULTIMO_PRECIO_NETO"]
    sobrestock = modelo[modelo["ESTADO_INVENTARIO"] == "SOBRESTOCK"].copy()
    sobrestock["EXCESO"] = (
        sobrestock["STOCK_ACTUAL"]
        - sobrestock["DEMANDA_MENSUAL"] * sobrestock["COBERTURA_OBJETIVO"]
    )
    sobrestock["VALOR_EXCESO"] = sobrestock["EXCESO"] * sobrestock["ULTIMO_PRECIO_NETO"]
    a_comprar = modelo[
        (modelo["COMPRA_AJUSTADA"] > 0) & modelo["ULTIMO_PRECIO_NETO"].notna()
    ]
    sin_costo = modelo[
        (modelo["ESTADO_INVENTARIO"].isin(["CRÍTICO", "COMPRAR"]))
        & modelo["ULTIMO_PRECIO_NETO"].isna()
    ]

    filas = [
        ("Fecha de análisis", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")),
        ("Período de demanda", f"{CONFIG['periodo_meses']} meses"),
        ("", ""),
        ("MÉTRICAS GENERALES", ""),
        ("Productos con demanda", f"{len(modelo):,}"),
        ("Demanda total (unidades)", f"{int(modelo['DEMANDA_TOTAL'].sum()):,}"),
        ("Demanda mensual (unidades)", f"{int(modelo['DEMANDA_MENSUAL'].sum()):,}"),
        ("", ""),
        ("CLASIFICACIÓN ABC", ""),
        ("Productos A", f"{(modelo['CLASIFICACION_ABC']=='A').sum():,}"),
        ("Productos B", f"{(modelo['CLASIFICACION_ABC']=='B').sum():,}"),
        ("Productos C", f"{(modelo['CLASIFICACION_ABC']=='C').sum():,}"),
        ("", ""),
        ("ESTADO DEL INVENTARIO", ""),
        ("Productos CRÍTICO", f"{(modelo['ESTADO_INVENTARIO']=='CRÍTICO').sum():,}"),
        ("Productos COMPRAR", f"{(modelo['ESTADO_INVENTARIO']=='COMPRAR').sum():,}"),
        ("Productos OK", f"{(modelo['ESTADO_INVENTARIO']=='OK').sum():,}"),
        ("Productos SOBRESTOCK", f"{(modelo['ESTADO_INVENTARIO']=='SOBRESTOCK').sum():,}"),
        ("", ""),
        ("INVERSIÓN SUGERIDA", ""),
        ("Productos a comprar este mes", f"{len(a_comprar):,}"),
        (
            "Inversión total sugerida (CLP)",
            f"${a_comprar['COSTO_COMPRA_AJUSTADA'].sum():,.0f}",
        ),
        ("", ""),
        ("CAPITAL INMOVILIZADO", ""),
        ("Valor stock total (CLP)", f"${modelo['VALOR_STOCK'].sum():,.0f}"),
        (
            "Valor exceso en sobrestock (CLP)",
            f"${sobrestock['VALOR_EXCESO'].sum():,.0f}",
        ),
        ("", ""),
        ("ALERTAS DE GESTIÓN", ""),
        (
            "Productos A en CRÍTICO",
            f"{((modelo['CLASIFICACION_ABC']=='A') & (modelo['ESTADO_INVENTARIO']=='CRÍTICO')).sum():,}",
        ),
        ("Productos sin costo registrado (CRÍTICO/COMPRAR)", f"{len(sin_costo):,}"),
        ("Solicitudes SIN ID (no catalogados)", f"{len(ranking_sin_id):,}"),
    ]

    for i, (concepto, valor) in enumerate(filas, start=3):
        c1 = ws.cell(row=i, column=1, value=concepto)
        c2 = ws.cell(row=i, column=2, value=valor)
        c1.font = FONT_NORMAL
        c2.font = FONT_NORMAL
        if concepto and not valor:  # Encabezado de sección
            c1.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            c1.fill = FILL_HEADER
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
            c1.alignment = ALIGN_LEFT

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 28


def hoja_modelo_completo(wb: Workbook, modelo: pd.DataFrame):
    ws = wb.create_sheet("2. Modelo Completo")
    cols = [
        "Id", "Marca", "Producto", "SKU", "Precio AILOO",
        "VENTAS_6M", "FALLIDAS_6M", "DEMANDA_TOTAL", "DEMANDA_MENSUAL",
        "PCT_PARTICIPACION", "PCT_ACUMULADO", "CLASIFICACION_ABC",
        "STOCK_ACTUAL", "EN_TRANSITO", "COBERTURA_MESES",
        "TIPO_ABASTECIMIENTO", "COBERTURA_OBJETIVO",
        "CANT_SUGERIDA_BASE", "ESTADO_INVENTARIO",
        "ULTIMO_PRECIO_NETO", "ULTIMA_FECHA_COMPRA", "ULTIMO_PROVEEDOR",
        "MESES_FINAL", "COMPRA_AJUSTADA", "COSTO_COMPRA_AJUSTADA",
        "OBSERVACION_COMPRA",
    ]
    rename = {
        "Id": "ID", "Marca": "MARCA", "Producto": "PRODUCTO", "SKU": "SKU",
        "Precio AILOO": "PRECIO AILOO",
        "VENTAS_6M": "VENTAS 6M", "FALLIDAS_6M": "FALLIDAS 6M",
        "DEMANDA_TOTAL": "DEMANDA TOTAL", "DEMANDA_MENSUAL": "DEMANDA MENSUAL",
        "PCT_PARTICIPACION": "% PARTICIPACION", "PCT_ACUMULADO": "% ACUMULADO",
        "CLASIFICACION_ABC": "ABC",
        "STOCK_ACTUAL": "STOCK ACTUAL", "EN_TRANSITO": "EN TRÁNSITO",
        "COBERTURA_MESES": "COBERTURA (MESES)",
        "TIPO_ABASTECIMIENTO": "TIPO ABASTECIMIENTO",
        "COBERTURA_OBJETIVO": "COBERTURA OBJETIVO",
        "CANT_SUGERIDA_BASE": "CANT SUGERIDA BASE",
        "ESTADO_INVENTARIO": "ESTADO INVENTARIO",
        "ULTIMO_PRECIO_NETO": "ULTIMO PRECIO NETO",
        "ULTIMA_FECHA_COMPRA": "ULTIMA FECHA COMPRA",
        "ULTIMO_PROVEEDOR": "ULTIMO PROVEEDOR",
        "MESES_FINAL": "MESES A COMPRAR",
        "COMPRA_AJUSTADA": "COMPRA AJUSTADA",
        "COSTO_COMPRA_AJUSTADA": "COSTO COMPRA AJUSTADA",
        "OBSERVACION_COMPRA": "OBSERVACION",
    }
    salida = modelo[cols].rename(columns=rename).copy()

    # Ordenar por prioridad: ABC + Estado + Demanda
    prio_estado = {"CRÍTICO": 1, "COMPRAR": 2, "OK": 3, "SOBRESTOCK": 4}
    prio_abc = {"A": 1, "B": 2, "C": 3}
    salida["_p1"] = salida["ABC"].map(prio_abc)
    salida["_p2"] = salida["ESTADO INVENTARIO"].map(prio_estado)
    salida = salida.sort_values(
        ["_p1", "_p2", "DEMANDA MENSUAL"], ascending=[True, True, False]
    ).drop(columns=["_p1", "_p2"])

    escribir_tabla(ws, salida)
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions


def hoja_prioridad(wb: Workbook, modelo: pd.DataFrame):
    ws = wb.create_sheet("3. Prioridad Compra")

    # Base común: productos que necesitan comprarse (CRÍTICO o COMPRAR) con cantidad >0
    base = modelo[
        (modelo["ESTADO_INVENTARIO"].isin(["CRÍTICO", "COMPRAR"]))
        & (modelo["COMPRA_AJUSTADA"] > 0)
    ].copy()

    cols_sel = [
        "Id", "Marca", "Producto", "CLASIFICACION_ABC", "ESTADO_INVENTARIO",
        "DEMANDA_MENSUAL", "STOCK_ACTUAL", "COBERTURA_MESES",
        "MESES_FINAL", "COMPRA_AJUSTADA", "Precio AILOO", "ULTIMO_PRECIO_NETO",
        "COSTO_COMPRA_AJUSTADA", "ULTIMO_PROVEEDOR",
    ]
    rename_cols = {
        "Id": "ID", "Marca": "MARCA", "Producto": "PRODUCTO",
        "CLASIFICACION_ABC": "ABC", "ESTADO_INVENTARIO": "ESTADO INVENTARIO",
        "DEMANDA_MENSUAL": "DEMANDA MENSUAL", "STOCK_ACTUAL": "STOCK ACTUAL",
        "COBERTURA_MESES": "COBERTURA (MESES)",
        "MESES_FINAL": "MESES A COMPRAR",
        "COMPRA_AJUSTADA": "COMPRA AJUSTADA",
        "Precio AILOO": "PRECIO AILOO",
        "ULTIMO_PRECIO_NETO": "ULTIMO PRECIO NETO",
        "COSTO_COMPRA_AJUSTADA": "COSTO COMPRA AJUSTADA",
        "ULTIMO_PROVEEDOR": "ULTIMO PROVEEDOR",
    }

    # ---- SECCIÓN 1: PRODUCTOS LISTOS PARA COMPRAR (con costo) ----
    df_con_costo = base[base["ULTIMO_PRECIO_NETO"].notna()][cols_sel].rename(columns=rename_cols)
    df_con_costo = df_con_costo.sort_values(
        ["ABC", "ESTADO INVENTARIO", "COSTO COMPRA AJUSTADA"],
        ascending=[True, True, False],
    )

    # ---- SECCIÓN 2: PRODUCTOS PENDIENTES DE COSTO ----
    df_sin_costo = base[base["ULTIMO_PRECIO_NETO"].isna()][cols_sel].rename(columns=rename_cols)
    df_sin_costo = df_sin_costo.sort_values(
        ["ABC", "ESTADO INVENTARIO", "DEMANDA MENSUAL"],
        ascending=[True, True, False],
    )

    inversion_total = df_con_costo["COSTO COMPRA AJUSTADA"].sum()

    # ============ ENCABEZADO SECCIÓN 1 ============
    fila = 1
    ws.cell(row=fila, column=1, value="SECCIÓN 1 — PRODUCTOS LISTOS PARA COMPRAR (con costo conocido)")
    ws.cell(row=fila, column=1).font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws.cell(row=fila, column=1).fill = PatternFill("solid", start_color="2E7D32")
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(cols_sel))
    ws.cell(row=fila, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 22

    fila += 1
    ws.cell(row=fila, column=1, value=f"Total: {len(df_con_costo):,} productos | Inversión sugerida: ${inversion_total:,.0f}")
    ws.cell(row=fila, column=1).font = Font(name="Arial", size=10, italic=True)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(cols_sel))

    # Headers sección 1
    fila += 2
    fila_header_1 = fila
    for j, col in enumerate(df_con_costo.columns, start=1):
        c = ws.cell(row=fila, column=j, value=col)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN

    # Datos sección 1
    for i, row in enumerate(df_con_costo.itertuples(index=False), start=fila + 1):
        for j, val in enumerate(row, start=1):
            escribir_fila_dato(ws, i, j, val, df_con_costo.columns[j - 1])
    fila += len(df_con_costo) + 1

    # ============ ENCABEZADO SECCIÓN 2 ============
    fila += 2  # espacio entre secciones
    ws.cell(row=fila, column=1, value=f"SECCIÓN 2 — PRODUCTOS PENDIENTES DE COSTO ({len(df_sin_costo):,} items con cantidad calculada pero sin monto)")
    ws.cell(row=fila, column=1).font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws.cell(row=fila, column=1).fill = PatternFill("solid", start_color="C00000")
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(cols_sel))
    ws.cell(row=fila, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 22

    fila += 1
    ws.cell(row=fila, column=1, value="⚠️ Estos productos REQUIEREN COMPRA pero falta levantar el costo con el proveedor. Cantidad calculada usando cobertura objetivo ABC (2.5/2/1 meses).")
    ws.cell(row=fila, column=1).font = Font(name="Arial", size=10, italic=True, color="C00000")
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(cols_sel))

    # Headers sección 2
    fila += 2
    for j, col in enumerate(df_sin_costo.columns, start=1):
        c = ws.cell(row=fila, column=j, value=col)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN

    # Datos sección 2
    for i, row in enumerate(df_sin_costo.itertuples(index=False), start=fila + 1):
        for j, val in enumerate(row, start=1):
            escribir_fila_dato(ws, i, j, val, df_sin_costo.columns[j - 1])

    # Anchos
    for j, col in enumerate(df_con_costo.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = ANCHOS.get(col, 12)

    ws.freeze_panes = f"D{fila_header_1 + 1}"


def hoja_sobrestock(wb: Workbook, modelo: pd.DataFrame):
    ws = wb.create_sheet("4. Sobrestock")
    df = modelo[modelo["ESTADO_INVENTARIO"] == "SOBRESTOCK"].copy()
    df["EXCESO_UNIDADES"] = (
        df["STOCK_ACTUAL"] - df["DEMANDA_MENSUAL"] * df["COBERTURA_OBJETIVO"]
    )
    df["VALOR_EXCESO"] = df["EXCESO_UNIDADES"] * df["ULTIMO_PRECIO_NETO"]
    df = df[
        ["Id", "Marca", "Producto", "CLASIFICACION_ABC", "STOCK_ACTUAL",
         "DEMANDA_MENSUAL", "COBERTURA_MESES", "COBERTURA_OBJETIVO",
         "EXCESO_UNIDADES", "ULTIMO_PRECIO_NETO", "Precio AILOO",
         "ULTIMO_PROVEEDOR", "VALOR_EXCESO"]
    ].rename(columns={
        "Id": "ID", "Marca": "MARCA", "Producto": "PRODUCTO",
        "CLASIFICACION_ABC": "ABC", "STOCK_ACTUAL": "STOCK",
        "DEMANDA_MENSUAL": "DEMANDA MENSUAL",
        "COBERTURA_MESES": "COBERTURA ACTUAL",
        "COBERTURA_OBJETIVO": "COBERTURA OBJ",
        "EXCESO_UNIDADES": "EXCESO UNIDADES",
        "ULTIMO_PRECIO_NETO": "COSTO UNITARIO",
        "Precio AILOO": "PRECIO AILOO",
        "ULTIMO_PROVEEDOR": "ULTIMO PROVEEDOR",
        "VALOR_EXCESO": "VALOR EXCESO",
    })
    df = df.sort_values("VALOR EXCESO", ascending=False)
    escribir_tabla(ws, df)
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions


def hoja_sin_costo(wb: Workbook, modelo: pd.DataFrame):
    ws = wb.create_sheet("5. Acción Costos")
    ws["A1"] = (
        "Productos en CRÍTICO o COMPRAR pero SIN costo registrado. "
        "Gestionar urgente con proveedor para habilitar compra."
    )
    ws["A1"].font = Font(name="Arial", size=10, italic=True, color="C00000")
    ws.merge_cells("A1:J1")

    df = modelo[
        (modelo["ESTADO_INVENTARIO"].isin(["CRÍTICO", "COMPRAR"]))
        & modelo["ULTIMO_PRECIO_NETO"].isna()
    ].copy()
    df = df.sort_values(
        ["CLASIFICACION_ABC", "DEMANDA_MENSUAL"], ascending=[True, False]
    )
    df = df[
        ["Id", "Marca", "Producto", "SKU", "CLASIFICACION_ABC",
         "DEMANDA_MENSUAL", "STOCK_ACTUAL", "COBERTURA_MESES",
         "ESTADO_INVENTARIO", "Precio AILOO"]
    ].rename(columns={
        "Id": "ID", "Marca": "MARCA", "Producto": "PRODUCTO",
        "CLASIFICACION_ABC": "ABC", "DEMANDA_MENSUAL": "DEMANDA MENSUAL",
        "STOCK_ACTUAL": "STOCK", "COBERTURA_MESES": "COBERTURA (MESES)",
        "ESTADO_INVENTARIO": "ESTADO", "Precio AILOO": "PRECIO AILOO",
    })
    escribir_tabla(ws, df, fila_inicio_header=3, fila_inicio_datos=4)
    ws.freeze_panes = "D4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(df.columns))}{len(df) + 3}"


def hoja_oportunidades_sin_id(wb: Workbook, ranking: pd.DataFrame):
    ws = wb.create_sheet("6. Oportunidades SIN ID")
    ws["A1"] = (
        f"TOP {CONFIG['excel']['max_filas_oportunidad_sin_id']} productos "
        "solicitados que NO están en el catálogo (oportunidad de ampliar surtido)"
    )
    ws["A1"].font = Font(name="Arial", size=10, italic=True, color="8B0000")
    ws.merge_cells("A1:B1")

    df = ranking.head(CONFIG["excel"]["max_filas_oportunidad_sin_id"]).copy()
    escribir_tabla(ws, df, fila_inicio_header=3, fila_inicio_datos=4)
    ws.freeze_panes = "A4"


# ============================================================================
# ENTRY POINT
# ============================================================================
def generar_excel(
    modelo: pd.DataFrame,
    ranking_sin_id: pd.DataFrame,
    output_path: Path,
    logger: logging.Logger,
):
    """Genera el archivo Excel completo con todas las hojas."""
    wb = Workbook()
    wb.remove(wb.active)

    hoja_resumen(wb, modelo, ranking_sin_id)
    hoja_modelo_completo(wb, modelo)
    hoja_prioridad(wb, modelo)
    hoja_sobrestock(wb, modelo)
    hoja_sin_costo(wb, modelo)
    hoja_oportunidades_sin_id(wb, ranking_sin_id)

    wb.save(output_path)
    logger.info(f"  ✓ Excel guardado: {output_path}")
    logger.info(f"  ✓ Tamaño: {output_path.stat().st_size / 1024:.1f} KB")
